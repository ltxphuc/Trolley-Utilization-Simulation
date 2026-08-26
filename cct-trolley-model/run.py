"""
RUN EVERYTHING
==============
    python run.py

Reads data/planppc.xlsx, runs every scenario, and writes into out/:

    dashboard.html   <- open this. Sliders, live numbers.
    results.xlsx     <- every scenario and every plan date, as a spreadsheet
    results.json     <- the same numbers, for anything that wants to read them
    sweep.json       <- what the dashboard's sliders read from

Nothing here decides anything. Every assumption is in assumptions.py.
"""
import json
import os
import time

import assumptions as A
import model as M
import scenarios as S
from dashboard import build_dashboard
import daytrace
from daytrace import build_trace
from model import fleet_required, load_days, longest_sustainable_cycle, run

OUT = 'out'


def summarise(result: dict) -> dict:
    """Strip the per-day detail and add the fleet arithmetic."""
    per_day = result.pop('per_day')
    tpd = result['trolleys_per_day']
    result['fleet_required_allowed_cycle'] = fleet_required(tpd, A.CYCLE_DAYS_ALLOWED)
    result['fleet_required_estimated_cycle'] = fleet_required(tpd, A.CYCLE_DAYS_ESTIMATED)
    result['fleet_gap_estimated_cycle'] = A.FLEET_SIZE - result['fleet_required_estimated_cycle']
    result['longest_sustainable_cycle_days'] = longest_sustainable_cycle(tpd)
    result['fits_in_fleet'] = result['fleet_required_estimated_cycle'] <= A.FLEET_SIZE
    return per_day


def main():
    os.makedirs(OUT, exist_ok=True)
    started = time.time()

    print('Reading', A.DATA_FILE)
    days, skipped = load_days(with_stats=True)
    print(f'  {len(days)} plan dates, '
          f'{sum(len(v) for v in days.values())} tables, '
          f'{sum(len(v) for v in days.values()) / len(days):.0f} per day\n')

    # ------------------------------------------------ headline scenarios
    print('HEADLINE SCENARIOS')
    print(f"{'':<32}{'util':>8}{'trolleys/day':>14}{'peak':>7}"
          f"{'fleet @' + str(A.CYCLE_DAYS_ESTIMATED) + 'd':>12}{'vs fleet':>11}")
    headline, headline_days = {}, {}
    for scenario in S.HEADLINE:
        r = run(scenario, days)
        headline_days[scenario.name] = summarise(r)
        headline[scenario.name] = r
        gap = r['fleet_gap_estimated_cycle']
        print(f"{scenario.name:<32}{r['utilization_pct']:>7.1f}%"
              f"{r['trolleys_per_day']:>14.0f}{r['peak_trolleys_in_a_day']:>7}"
              f"{r['fleet_required_estimated_cycle']:>12,.0f}"
              f"{gap:>+11,.0f}{'' if gap >= 0 else '  SHORT'}")

    base, best = S.HEADLINE[0].name, S.HEADLINE[-1].name
    print(f"\n  fleet = {A.FLEET_SIZE:,} trolleys · "
          f"cycle {A.CYCLE_DAYS_ALLOWED}d allowed, ~{A.CYCLE_DAYS_ESTIMATED}d estimated")
    print(f"  {base} sustains a "
          f"{headline[base]['longest_sustainable_cycle_days']:.1f}-day cycle; "
          f"{best} sustains "
          f"{headline[best]['longest_sustainable_cycle_days']:.1f} days.\n")

    # ------------------------------------------- same WS vs any WS
    print('WIP HOLD — WHERE MAY A PARKED TROLLEY GO?')
    print(f"{'':<32}{'util':>8}{'trolleys/day':>14}{'avg wait':>11}")
    variants = {}
    for scenario in S.HOLD_VARIANTS:
        r = run(scenario, days)
        summarise(r)
        variants[scenario.name] = r
        print(f"{scenario.name:<32}{r['utilization_pct']:>7.1f}%"
              f"{r['trolleys_per_day']:>14.0f}{r['avg_wait_hours']:>10.1f}h")
    print()

    # ------------------------------------------------------- the sweep
    combos = S.sweep_scenarios()
    print(f'SWEEP — {len(combos)} combinations for the dashboard sliders')
    sweep = {}
    for i, (key, scenario) in enumerate(combos.items(), 1):
        r = run(scenario, days)
        summarise(r)
        sweep[key] = {
            'utilization_pct': round(r['utilization_pct'], 2),
            'trolleys_per_day': round(r['trolleys_per_day'], 1),
            'peak': r['peak_trolleys_in_a_day'],
            'pieces_per_trolley': round(r['pieces_per_trolley'], 0),
            'avg_wait_hours': round(r['avg_wait_hours'], 2),
            'spreading_blocked_pct': round(r['spreading_blocked_pct'], 1),
            'cutting_blocked_pct': round(r['cutting_blocked_pct'], 1),
            'blocked_by_space': r['blocked_by_space'],
        }
        if i % 10 == 0 or i == len(combos):
            print(f'  {i}/{len(combos)}')
    print()

    # ------------------------------------------------- per-day detail
    # What the Simulation tab draws: every table of every day, the sequence
    # it was given, and what Python makes of the loading — so the page can
    # check its own arithmetic against the model in front of the reader.
    print('TRACE — per-day detail for the Simulation tab')
    trace = build_trace(days, skipped)
    print(f'  {len(trace["dates"])} days, '
          f'{sum(len(d["tasks"]) for d in trace["days"].values())} tables\n')

    payload = {
        'generated_seconds': round(time.time() - started, 1),
        'plan_dates': len(days),
        'tables': sum(len(v) for v in days.values()),
        'assumptions': {
            'fleet_size': A.FLEET_SIZE,
            'cycle_days_allowed': A.CYCLE_DAYS_ALLOWED,
            'cycle_days_estimated': A.CYCLE_DAYS_ESTIMATED,
            'trolley_compartments': A.TROLLEY_COMPARTMENTS,
            'compartment_cap': A.COMPARTMENT_CAP,
            'max_batches_per_trolley': A.MAX_BATCHES_PER_TROLLEY,
            'trolley_bays': A.TROLLEY_BAYS,
            'wip_hold_minutes': A.WIP_HOLD_MINUTES,
            'wip_hold_any_workstation': A.WIP_HOLD_ANY_WORKSTATION,
            'max_trolleys_per_combine': A.MAX_TROLLEYS_PER_COMBINE,
            'relax_hours': A.RELAX_HOURS,
            'spreading_table_length_m': A.SPREADING_TABLE_LENGTH_M,
            'shift_start_minute': A.SHIFT_START_MINUTE,
            'film_size_by_garment_size': A.FILM_SIZE_BY_GARMENT_SIZE,
        },
        'cut_groups': A.CUT_GROUPS,
        'night_shift_spreaders': sorted(A.NIGHT_SHIFT_SPREADERS),
        'night_shift_cutters': sorted(A.NIGHT_SHIFT_CUTTERS),
        # what the Gantt draws its shift bands and break gaps from
        'spreader_shift': {m: M.SPREADER_SHIFT[m] for m in M.SPREADERS},
        'cutter_shift': {c: M.CUTTER_SHIFT[c] for c in A.CUT_GROUPS},
        'machine_minutes': M.MACHINE_MINUTES,
        'shift_end': M.SHIFT_END,
        'day_shift_end': A.DAY_SHIFT[1],
        'table_lengths_m': [f'{m:g}' for m in daytrace.TABLE_LENGTHS_M],
        'status_colors': A.TASK_STATUS_COLORS,
        'status_not_simulated': A.TASK_STATUS_NOT_SIMULATED,
        'headline': headline,
        'hold_variants': variants,
        'sweep': sweep,
        'sweep_axes': {
            'hold_minutes': S.SWEEP_HOLD_MINUTES,
            'max_batches': S.SWEEP_MAX_BATCHES,
        },
        'per_day': headline_days,
    }

    with open(f'{OUT}/results.json', 'w') as fh:
        json.dump(payload, fh, indent=1, default=str)
    with open(f'{OUT}/sweep.json', 'w') as fh:
        json.dump({'sweep': sweep, 'assumptions': payload['assumptions']}, fh, indent=1)
    print(f'wrote {OUT}/results.json  {OUT}/sweep.json')

    write_excel(payload)
    # The trace is handed to the dashboard directly rather than through
    # results.json — it is per-table detail for the page to draw, and it would
    # bury the scenario numbers that file exists to carry.
    build_dashboard(payload, trace, f'{OUT}/dashboard.html')
    print(f'wrote {OUT}/dashboard.html')
    print(f'\ndone in {time.time() - started:.0f}s — open {OUT}/dashboard.html')


def write_excel(payload: dict):
    """One workbook: scenario summary, per-day detail, and the raw sweep."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        print('  (openpyxl missing — skipping Excel)')
        return

    wb = openpyxl.Workbook()
    head_font = Font(bold=True, color='FFFFFF')
    head_fill = PatternFill('solid', fgColor='1E2A38')

    def sheet(title, headers, rows, widths=None):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for c in ws[1]:
            c.font, c.fill = head_font, head_fill
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for r in rows:
            ws.append(r)
        for i, w in enumerate(widths or [], start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'
        return ws

    # -- summary
    rows = []
    for name, r in payload['headline'].items():
        rows.append([
            name, round(r['utilization_pct'], 1), round(r['trolleys_per_day'], 0),
            r['peak_trolleys_in_a_day'], round(r['pieces_per_trolley'], 0),
            round(r['fleet_required_allowed_cycle'], 0),
            round(r['fleet_required_estimated_cycle'], 0),
            round(r['fleet_gap_estimated_cycle'], 0),
            round(r['longest_sustainable_cycle_days'], 1),
            'yes' if r['fits_in_fleet'] else 'NO',
            round(r['spreading_blocked_pct'], 1), round(r['cutting_blocked_pct'], 1),
        ])
    sheet('Scenarios',
          ['Scenario', 'Utilization %', 'Trolleys/day', 'Peak day', 'Pieces/trolley',
           f"Fleet needed @{payload['assumptions']['cycle_days_allowed']}d",
           f"Fleet needed @{payload['assumptions']['cycle_days_estimated']}d",
           'Fleet spare', 'Cycle days the fleet covers', 'Fits in fleet?',
           'Spreading blocked %', 'Cutting blocked %'],
          rows, [30, 13, 12, 10, 13, 13, 13, 11, 14, 11, 13, 13])

    # -- hold variants
    sheet('WIP hold variants',
          ['Setting', 'Utilization %', 'Trolleys/day', 'Avg wait (h)',
           f"Fleet needed @{payload['assumptions']['cycle_days_estimated']}d", 'Fits?'],
          [[n, round(r['utilization_pct'], 1), round(r['trolleys_per_day'], 0),
            round(r['avg_wait_hours'], 1),
            round(r['fleet_required_estimated_cycle'], 0),
            'yes' if r['fits_in_fleet'] else 'NO']
           for n, r in payload['hold_variants'].items()],
          [32, 13, 12, 12, 14, 8])

    # -- per day
    rows = []
    for name, per_day in payload['per_day'].items():
        for date, d in sorted(per_day.items()):
            rows.append([name, date, d['tasks'], d['combines'], d['compartments'],
                         d['trolleys'], round(d['trolley_utilization_pct'], 1),
                         d['pieces'], round(d['spreading_blocked_pct'], 1),
                         round(d['cutting_blocked_pct'], 1),
                         round(d['spreading_makespan_hours'], 1)])
    sheet('Per plan date',
          ['Scenario', 'Plan date', 'Tables', 'Combines', 'Compartments', 'Trolleys',
           'Utilization %', 'Pieces', 'Spreading blocked %', 'Cutting blocked %',
           'Spreading makespan (h)'],
          rows, [30, 12, 8, 10, 13, 9, 13, 10, 17, 17, 19])

    # -- sweep
    rows = []
    for key, r in payload['sweep'].items():
        mo, batches, hold, any_ws = key.split('|')
        rows.append(['yes' if mo == '1' else 'no', int(batches), int(hold),
                     'any' if any_ws == '1' else 'same',
                     r['utilization_pct'], r['trolleys_per_day'], r['peak'],
                     r['avg_wait_hours']])
    sheet('Sweep (all settings)',
          ['Same-MO sequencing', 'Max batches/trolley', 'WIP hold (min)',
           'Return to workstation', 'Utilization %', 'Trolleys/day', 'Peak',
           'Avg wait (h)'],
          rows, [18, 18, 15, 19, 13, 12, 8, 12])

    # -- assumptions, so the workbook is self-describing
    sheet('Assumptions', ['Setting', 'Value'],
          [[k.replace('_', ' '), v] for k, v in payload['assumptions'].items()],
          [34, 14])

    wb.remove(wb['Sheet'])
    wb.save(f'{OUT}/results.xlsx')
    print(f'wrote {OUT}/results.xlsx')


if __name__ == '__main__':
    main()
